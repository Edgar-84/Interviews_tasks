import string
from openpyxl import load_workbook, Workbook


class ExcelModule:

    COLUMS_POSITION = {key: value for key, value in enumerate(string.ascii_uppercase)}
    
    @staticmethod
    def get_maximum_rows(sheet_object: Workbook) -> int:
        rows = 0
        for _, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1

        return rows

    @staticmethod
    def get_name_tabs(path_excel: str, number_list: int = 0) -> list:
        wb = load_workbook(path_excel)
        sheet = wb.worksheets[number_list]
        result = [sheet[1][column].value for column in range(sheet.max_column)]
        
        return result

    @classmethod
    def write_name_tabs(cls, path_file: str, data: list):
        wb = Workbook()
        sheet = wb.active

        for number, value in enumerate(data):
            sheet[f'{cls.COLUMS_POSITION[number]}1'] = value

        wb.save(path_file)


    @classmethod
    def get_data(cls, path_excel: str, name_tabs: bool = True) -> list[list[str]]:
        wb = load_workbook(path_excel)
        sheet = wb.active 
        first = 1 if name_tabs else 2
        excel_data = []
        result = []

        for row in range(first, cls.get_maximum_rows(sheet) + 1):
            for column in range(sheet.max_column):
                excel_data.append(sheet[row][column].value)

            result.append(excel_data)
            excel_data = []
        
        return result
    

    @classmethod
    def write_data(cls, path_file: str, data: list, column: int):     
        wb = load_workbook(path_file)
        sheet = wb.active

        for row, value in enumerate(data):
            for num in range(column):
                sheet[f'{cls.COLUMS_POSITION[num]}{row + 2}'] = value[num]

        wb.save(path_file)

